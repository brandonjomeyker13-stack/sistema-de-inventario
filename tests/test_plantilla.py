"""
La plantilla de Excel que el tendero descarga.

La genera el backend para que los encabezados salgan de la MISMA clase con
la que el importador lee el archivo. Estas pruebas comprueban justo eso: que
lo que se entrega y lo que se lee no puedan separarse.
"""

from io import BytesIO

from openpyxl import load_workbook

from app.core import plantilla
from app.schemas.importacion import FilaImportar


def _abrir():
    return load_workbook(BytesIO(plantilla.generar()))


def test_los_encabezados_son_exactamente_los_que_lee_el_importador():
    """La prueba que evita el fallo silencioso: si algún día se agrega una
    columna a FilaImportar y la plantilla no la trae, el importador
    rechazaría archivos sin que nadie entienda por qué."""
    hoja = _abrir()["Productos"]
    encabezados = [c.value for c in hoja[1]]

    assert encabezados == list(FilaImportar.model_fields.keys())


def test_la_hoja_de_datos_viene_vacia():
    """Sin filas de ejemplo. Si las trajera y el tendero olvidara borrarlas,
    se importarían como productos reales: tendría un "Arroz Diana" inventado
    con existencias que no tiene."""
    hoja = _abrir()["Productos"]
    assert hoja.max_row == 1


def test_trae_una_hoja_de_instrucciones():
    libro = _abrir()
    assert "Instrucciones" in libro.sheetnames
    # La primera hoja debe ser la de llenar: es la que abre Excel y la que
    # va a leer el frontend.
    assert libro.sheetnames[0] == "Productos"


def test_las_instrucciones_explican_todas_las_columnas():
    hoja = _abrir()["Instrucciones"]
    texto = " ".join(str(c.value) for fila in hoja.iter_rows() for c in fila if c.value)

    for columna in FilaImportar.model_fields:
        assert columna in texto, f"falta explicar la columna '{columna}'"


def test_las_instrucciones_avisan_de_lo_que_mas_se_equivoca():
    hoja = _abrir()["Instrucciones"]
    texto = " ".join(str(c.value) for fila in hoja.iter_rows() for c in fila if c.value)

    # Invertir precio y costo es el error de dedo más común, y el que deja
    # un producto que pierde plata en cada venta.
    assert "invertirlo" in texto or "invertir" in texto
    # Y que los códigos de barras no van aquí.
    assert "código" in texto.lower() or "codigo" in texto.lower()


def test_el_ejemplo_incluye_un_servicio():
    """Una papelería no puede representar sus fotocopias sin esa columna, y
    es el caso que motivó todo. Verlo en el ejemplo lo enseña sin explicar."""
    hoja = _abrir()["Instrucciones"]
    texto = " ".join(str(c.value) for fila in hoja.iter_rows() for c in fila if c.value)
    assert "Fotocopia" in texto


def test_el_archivo_es_un_xlsx_valido_y_pequeno():
    contenido = plantilla.generar()
    assert contenido[:2] == b"PK"          # firma de un zip, que es lo que es un xlsx
    assert len(contenido) < 100_000        # unos pocos KB: se manda por WhatsApp


def test_los_precios_del_ejemplo_estan_en_formato_colombiano():
    """El ejemplo enseña sin explicar: si muestra "2.800", el tendero escribe
    así y el backend lo entiende."""
    hoja = _abrir()["Instrucciones"]
    valores = [str(c.value) for fila in hoja.iter_rows() for c in fila if c.value]
    assert any("2.800" in v for v in valores)
