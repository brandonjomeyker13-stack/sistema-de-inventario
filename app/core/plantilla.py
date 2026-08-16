"""
app/core/plantilla.py — Genera el Excel que el tendero descarga y llena.

POR QUÉ LA GENERA EL BACKEND Y NO EL FRONTEND

Los encabezados salen de `FilaImportar`, que es la MISMA clase con la que el
importador lee el archivo. Así es imposible que se desincronicen: el día que
se agregue o se renombre una columna, la plantilla cambia sola.

Si el frontend los escribiera a mano, el día que cambie una columna la
plantilla seguiría entregando la anterior y el importador rechazaría todas
las filas sin que nadie entendiera por qué.

Y hay una ventaja práctica: al ser un archivo del servidor, un vendedor
puede mandárselo al tendero por WhatsApp antes de la visita, sin que nadie
tenga que abrir la aplicación.

LA HOJA DE EJEMPLOS VA APARTE

La hoja donde se escribe queda con SOLO los títulos. Los ejemplos van en una
segunda hoja de instrucciones. Si los ejemplos estuvieran en la hoja de
datos y el tendero olvidara borrarlos, se importarían como productos reales:
tendría un "Arroz Blanco" inventado en su inventario, con existencias que no
tiene.
"""

from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from app.schemas.importacion import MAX_FILAS, FilaImportar

NOMBRE_ARCHIVO = "plantilla-inventario-stocktraking.xlsx"

# El orden y los nombres salen del schema del importador. Ver el docstring.
COLUMNAS = list(FilaImportar.model_fields.keys())

# Lo que se explica de cada columna en la hoja de instrucciones.
AYUDA = {
    "nombre": (
        "OBLIGATORIO",
        "Como lo llamas tú. Si un producto con ese mismo nombre ya está en el "
        "sistema, se actualiza en vez de crearse otro. No importan las mayúsculas "
        "ni las tildes: \"Café\", \"cafe\" y \"CAFÉ\" son el mismo producto.",
    ),
    "cantidad": (
        "OBLIGATORIO",
        "Cuántas unidades tienes AHORA. Sin decimales: 12, no 12,5. "
        "Si es un servicio, escribe 0.",
    ),
    "precio": (
        "OBLIGATORIO",
        "A cuánto lo VENDES. Puedes escribirlo como quieras: 1500, 1.500 o $1.500.",
    ),
    "cuanto_costo": (
        "OBLIGATORIO",
        "A cuánto te cuesta COMPRARLO. Es lo que permite saber cuánto ganas. "
        "Ojo de no invertirlo con el precio.",
    ),
    "categoria": (
        "Opcional",
        "Para agrupar (Granos, Aseo, Bebidas). Si la dejas vacía el producto "
        "queda sin categoría y se le puede poner después. Si escribes una que "
        "no existe, se crea.",
    ),
    "es_servicio": (
        "Opcional",
        "Escribe \"si\" cuando NO lleva inventario: fotocopias, impresiones, "
        "plastificado, anillado. Esos no se acaban. Déjala vacía para un "
        "producto normal.",
    ),
}

EJEMPLOS = [
    ["Arroz Diana 500g", 24, "2.800", "2.200", "Granos", ""],
    ["Cuaderno argollado 100 hojas", 15, "6.500", "4.800", "Papelería", ""],
    ["Fotocopia", 0, "200", "50", "Servicios", "si"],
    ["Gaseosa personal", 36, "2.500", "1.900", "Bebidas", ""],
]

NOTAS = [
    "No borres ni cambies la primera fila (la de los títulos).",
    "Escribe un producto por fila, empezando en la fila 2.",
    f"Máximo {MAX_FILAS:,} productos por archivo.".replace(",", "."),
    "Los precios se pueden escribir con puntos y con el signo de peso: "
    "1500, 1.500 y $1.500 se leen igual.",
    "Los códigos de barras NO van en esta plantilla. Los productos se crean "
    "sin código y después se los pones escaneándolos con el celular.",
    "Antes de guardar nada, el sistema te muestra qué va a crear y qué va a "
    "cambiar, para que lo revises.",
    "Si una fila tiene un problema, no se importa NADA: te decimos qué filas "
    "corregir y vuelves a subir el archivo completo, sin miedo a duplicar.",
]

_AZUL = "1F4E79"
_GRIS = "F2F2F2"


def _hoja_de_datos(libro: Workbook) -> None:
    """La hoja que el tendero llena. Solo los títulos, sin ejemplos."""
    hoja = libro.active
    hoja.title = "Productos"

    for indice, columna in enumerate(COLUMNAS, start=1):
        celda = hoja.cell(row=1, column=indice, value=columna)
        celda.font = Font(bold=True, color="FFFFFF", size=11)
        celda.fill = PatternFill("solid", fgColor=_AZUL)
        celda.alignment = Alignment(horizontal="center", vertical="center")
        # Ancho generoso en el nombre, que es el campo largo.
        hoja.column_dimensions[get_column_letter(indice)].width = (
            34 if columna == "nombre" else 16
        )

    hoja.row_dimensions[1].height = 24
    # Congela la fila de títulos: con trescientos productos, saber en qué
    # columna se está sin subir hasta arriba importa.
    hoja.freeze_panes = "A2"


def _hoja_de_instrucciones(libro: Workbook) -> None:
    hoja = libro.create_sheet("Instrucciones")
    hoja.column_dimensions["A"].width = 18
    hoja.column_dimensions["B"].width = 14
    hoja.column_dimensions["C"].width = 82

    fila = 1

    def titulo(texto: str) -> None:
        nonlocal fila
        celda = hoja.cell(row=fila, column=1, value=texto)
        celda.font = Font(bold=True, size=14, color=_AZUL)
        fila += 2

    titulo("Cómo llenar esta plantilla")

    for indice, etiqueta in enumerate(("Columna", "¿Obligatoria?", "Qué se escribe"), start=1):
        celda = hoja.cell(row=fila, column=indice, value=etiqueta)
        celda.font = Font(bold=True, color="FFFFFF")
        celda.fill = PatternFill("solid", fgColor=_AZUL)
    fila += 1

    for columna in COLUMNAS:
        obligatoria, explicacion = AYUDA.get(columna, ("", ""))
        hoja.cell(row=fila, column=1, value=columna).font = Font(bold=True)
        hoja.cell(row=fila, column=2, value=obligatoria)
        celda = hoja.cell(row=fila, column=3, value=explicacion)
        celda.alignment = Alignment(wrap_text=True, vertical="top")
        hoja.row_dimensions[fila].height = 46
        fila += 1

    fila += 1
    titulo("Ejemplo")
    hoja.cell(row=fila - 1, column=3,
              value="Así se vería la hoja «Productos» llena. Estas filas NO se importan: "
                    "están aquí solo de muestra.").font = Font(italic=True, size=10)
    fila += 1

    for indice, columna in enumerate(COLUMNAS, start=1):
        celda = hoja.cell(row=fila, column=indice, value=columna)
        celda.font = Font(bold=True)
        celda.fill = PatternFill("solid", fgColor=_GRIS)
    fila += 1

    for ejemplo in EJEMPLOS:
        for indice, valor in enumerate(ejemplo, start=1):
            hoja.cell(row=fila, column=indice, value=valor)
        fila += 1

    fila += 2
    titulo("Cosas que conviene saber")
    for nota in NOTAS:
        celda = hoja.cell(row=fila, column=1, value=f"•  {nota}")
        celda.alignment = Alignment(wrap_text=True, vertical="top")
        hoja.merge_cells(start_row=fila, start_column=1, end_row=fila, end_column=3)
        hoja.row_dimensions[fila].height = 30
        fila += 1


def generar() -> bytes:
    """El archivo .xlsx listo para descargar.

    Se devuelve en memoria y no se guarda en disco: el contenido es siempre
    el mismo y depende solo del código, así que escribirlo en el disco de
    Render (que es efímero) solo añadiría un sitio donde quedarse
    desactualizado.
    """
    libro = Workbook()
    _hoja_de_datos(libro)
    _hoja_de_instrucciones(libro)

    memoria = BytesIO()
    libro.save(memoria)
    return memoria.getvalue()
